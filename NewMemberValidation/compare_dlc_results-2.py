"""
DLC Results Comparison Script
==============================
Compares DeepLabCut raw tracking outputs between two models (yours vs overseer's).
Correlates frame-by-frame body part coordinates and speed columns.

USAGE
-----
Set the two folder paths in the CONFIG section below, then run:
    python compare_dlc_results.py

OUTPUT
------
One Excel file: dlc_comparison_results.xlsx
  - One sheet per matched animal (frame-by-frame data: mine | hers | diff)
  - A Summary sheet at the front with Pearson r, MAE, and mean diff per animal per column
"""

import os
import re
import glob
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
#  CONFIG  ← edit these paths before running
# ─────────────────────────────────────────────────────────────

MY_FOLDER       = "/Users/acanahuati/Downloads/ProjectForAri/ari_dlc_csv"
OVERSEER_FOLDER = "/Users/acanahuati/Downloads/ProjectForAri/ellie_dlc_csv"

OUTPUT_FILE = "dlc_comparison_results.xlsx"

# Raw body part columns to correlate
TRACKING_COLS = [
    "Nose",           # Nose X coordinate
    "Nose.1",         # Nose Y coordinate
    "Tail_base",      # Tail base X coordinate
    "Tail_base.1",    # Tail base Y coordinate
    "Tailbase_Speed",
    "Nose_Speed",
]

# Human-readable labels for the summary (matches TRACKING_COLS order)
COL_LABELS = {
    "Nose":            "Nose X",
    "Nose.1":          "Nose Y",
    "Tail_base":       "Tail Base X",
    "Tail_base.1":     "Tail Base Y",
    "Tailbase_Speed":  "Tailbase Speed",
    "Nose_Speed":      "Nose Speed",
}

# ─────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────

def extract_animal_id(filename):
    basename = os.path.basename(filename)
    for ext in [".csv", ".numbers", "_filtered", "-velocity_data"]:
        basename = basename.replace(ext, "")
    match = re.match(r"(\d+[._]\d+)", basename)
    return match.group(1).replace(".", "_") if match else basename


def style_header(ws, row, n_cols, fill_hex="2F5496"):
    fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
    font = Font(bold=True, color="FFFFFF", name="Arial")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def style_subheader(ws, row, n_cols, fill_hex="D9E1F2"):
    fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
    font = Font(bold=True, color="1F3864", name="Arial")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def color_corr_cell(cell, r):
    """Green if r >= 0.9, orange if 0.7-0.9, red if < 0.7."""
    if pd.isna(r):
        return
    if r >= 0.9:
        cell.font = Font(bold=True, color="375623", name="Arial")
        cell.fill = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    elif r >= 0.7:
        cell.font = Font(bold=True, color="7F4D00", name="Arial")
        cell.fill = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
    else:
        cell.font = Font(bold=True, color="9C0006", name="Arial")
        cell.fill = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")


# ─────────────────────────────────────────────────────────────
#  FILE MATCHING
# ─────────────────────────────────────────────────────────────

def match_tracking_files(my_folder, her_folder):
    my_files  = {extract_animal_id(f): f for f in glob.glob(os.path.join(my_folder,  "*.csv"))}
    her_files = {extract_animal_id(f): f for f in glob.glob(os.path.join(her_folder, "*.csv"))}

    matched, only_mine, only_hers = [], [], []
    for aid in sorted(set(my_files) | set(her_files)):
        if aid in my_files and aid in her_files:
            matched.append((aid, my_files[aid], her_files[aid]))
        elif aid in my_files:
            only_mine.append(aid)
        else:
            only_hers.append(aid)

    if only_mine:
        print(f"[INFO] Only in YOUR folder (no match): {only_mine}")
    if only_hers:
        print(f"[INFO] Only in OVERSEER folder (no match): {only_hers}")
    return matched


def load_tracking_csv(path):
    df = pd.read_csv(path)
    if "Second" not in df.columns:
        df = df.rename(columns={df.columns[0]: "Second"})
    df["Second"] = pd.to_numeric(df["Second"], errors="coerce")
    df = df.dropna(subset=["Second"]).reset_index(drop=True)
    return df


# ─────────────────────────────────────────────────────────────
#  TRACKING SHEET  (frame-by-frame)
# ─────────────────────────────────────────────────────────────

def build_tracking_sheet(ws, my_df, her_df, animal_id):
    cols_present = [c for c in TRACKING_COLS if c in my_df.columns and c in her_df.columns]

    merged = (
        my_df[["Second"] + cols_present]
        .merge(her_df[["Second"] + cols_present],
               on="Second", suffixes=("_mine", "_hers"), how="outer")
        .sort_values("Second")
        .reset_index(drop=True)
    )

    # ── Title row ──
    title = f"Animal {animal_id} — Frame-by-Frame Comparison"
    ws.append([title])
    ws.cell(1, 1).font = Font(bold=True, size=12, name="Arial", color="1F3864")
    end_col = 1 + len(cols_present) * 3
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)

    # ── Column group header (Mine | Hers | Diff per body part) ──
    group_row = ["Second"]
    for c in cols_present:
        label = COL_LABELS.get(c, c)
        group_row += [f"<-- {label} -->", "", ""]
    ws.append(group_row)
    style_header(ws, 2, len(group_row))

    # ── Sub-header ──
    sub_row = ["Second (s)"]
    for _ in cols_present:
        sub_row += ["Mine", "Hers", "Diff (mine-hers)"]
    ws.append(sub_row)
    style_subheader(ws, 3, len(sub_row))

    # ── Data rows ──
    for _, row in merged.iterrows():
        data = [row["Second"]]
        for c in cols_present:
            mine = row.get(f"{c}_mine", np.nan)
            hers = row.get(f"{c}_hers", np.nan)
            diff = (mine - hers) if not (pd.isna(mine) or pd.isna(hers)) else np.nan
            data += [
                round(float(mine), 4) if not pd.isna(mine) else "",
                round(float(hers), 4) if not pd.isna(hers) else "",
                round(float(diff), 4) if not pd.isna(diff) else "",
            ]
        ws.append(data)

    for r in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in r:
            cell.font = Font(name="Arial", size=9)
            cell.border = thin_border()

    autofit(ws)
    return _tracking_stats(merged, cols_present, animal_id)


def _tracking_stats(merged, cols, animal_id):
    rows = []
    for c in cols:
        mine = pd.to_numeric(merged.get(f"{c}_mine"), errors="coerce").values
        hers = pd.to_numeric(merged.get(f"{c}_hers"), errors="coerce").values
        mask = ~(np.isnan(mine) | np.isnan(hers))
        if mask.sum() < 2:
            corr = mae = mean_diff = np.nan
        else:
            corr      = float(np.corrcoef(mine[mask], hers[mask])[0, 1])
            mae       = float(np.mean(np.abs(mine[mask] - hers[mask])))
            mean_diff = float(np.mean(mine[mask] - hers[mask]))
        rows.append({
            "Animal":                animal_id,
            "Column":                COL_LABELS.get(c, c),
            "N_frames":              int(mask.sum()),
            "Pearson_r":             round(corr, 4)      if not np.isnan(corr)      else "N/A",
            "MAE":                   round(mae, 4)        if not np.isnan(mae)        else "N/A",
            "Mean_Diff (mine-hers)": round(mean_diff, 4) if not np.isnan(mean_diff) else "N/A",
        })
    return rows


# ─────────────────────────────────────────────────────────────
#  SUMMARY SHEET
# ─────────────────────────────────────────────────────────────

def build_summary_sheet(ws, all_stats):
    ws.append(["DLC Raw Tracking Correlation Summary"])
    ws.cell(1, 1).font = Font(bold=True, size=13, name="Arial", color="1F3864")
    ws.merge_cells("A1:F1")

    ws.append(["Pearson r colour key:  Green >= 0.90 (strong)   Orange 0.70-0.89 (moderate)   Red < 0.70 (weak)"])
    ws.cell(2, 1).font = Font(italic=True, size=9, name="Arial", color="595959")
    ws.merge_cells("A2:F2")

    ws.append([])  # blank spacer

    if not all_stats:
        ws.append(["No data available — check that MY_FOLDER and OVERSEER_FOLDER are set correctly."])
        return

    df = pd.DataFrame(all_stats)
    cols = list(df.columns)

    ws.append(cols)
    style_header(ws, 4, len(cols))

    for _, row in df.iterrows():
        ws.append(list(row))

    # Colour-code the Pearson_r column
    r_col_idx = cols.index("Pearson_r") + 1
    for excel_row in range(5, ws.max_row + 1):
        cell = ws.cell(excel_row, r_col_idx)
        try:
            color_corr_cell(cell, float(cell.value))
        except (TypeError, ValueError):
            pass

    for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in r:
            if not cell.font or not cell.font.bold:
                cell.font = Font(name="Arial")
            cell.border = thin_border()

    autofit(ws)


# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)
    all_stats = []

    if not (os.path.isdir(MY_FOLDER) and os.path.isdir(OVERSEER_FOLDER)):
        print(f"[ERROR] One or both folders not found.")
        print(f"  MY_FOLDER:       {MY_FOLDER}")
        print(f"  OVERSEER_FOLDER: {OVERSEER_FOLDER}")
        return

    matched = match_tracking_files(MY_FOLDER, OVERSEER_FOLDER)

    if not matched:
        print("[ERROR] No matching animal IDs found between the two folders.")
        return

    for animal_id, my_path, her_path in matched:
        print(f"[INFO] Comparing: {animal_id}")
        my_df  = load_tracking_csv(my_path)
        her_df = load_tracking_csv(her_path)
        ws = wb.create_sheet(f"Track_{animal_id}"[:31])
        stats = build_tracking_sheet(ws, my_df, her_df, animal_id)
        all_stats.extend(stats)

    ws_summary = wb.create_sheet("Summary", 0)
    build_summary_sheet(ws_summary, all_stats)

    wb.save(OUTPUT_FILE)
    print(f"\nDone! Results saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
