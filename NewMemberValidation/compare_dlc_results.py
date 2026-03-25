"""
DLC Results Comparison Script
==============================
Compares DeepLabCut outputs between two models (yours vs overseer's).

Handles two types of comparison:
  1. Behaviour summary CSVs  (scan counts, freezing counts)
  2. Raw DLC tracking CSVs   (velocity / coordinate files)

USAGE
-----
Set the four paths in the CONFIG section below, then run:
    python compare_dlc_results.py

OUTPUT
------
One Excel file: dlc_comparison_results.xlsx
  - Sheet per behaviour metric (e.g. Scanning_Comparison, Freezing_Comparison)
  - Sheet per matched tracking file pair
  - A Summary sheet with correlation and MAE for every metric
"""

import os
import re
import glob
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
#  CONFIG  ← edit these paths before running
# ─────────────────────────────────────────────────────────────

MY_FOLDER        = "/Users/acanahuati/Downloads/ProjectForAri/ari_dlc_csv"       # folder with YOUR DLC CSVs
OVERSEER_FOLDER  = "/Users/acanahuati/Downloads/ProjectForAri/ellie_dlc_csv" # folder with HER DLC CSVs

# Behaviour summary CSVs (produced by rearing/scanning scripts).
# Set to None if you don't have an overseer equivalent yet.
MY_SCAN_CSV      = "/Users/acanahuati/Downloads/ProjectForAri/ari_script_csv/Hab_scan.csv"
MY_FREEZE_CSV    = "/Users/acanahuati/Downloads/ProjectForAri/ari_script_csv/Hab_freezing.csv"
OVERSEER_SCAN_CSV   = "/Users/acanahuati/Downloads/ProjectForAri/ellie_script_csv/Hab_scan.csv"
OVERSEER_FREEZE_CSV = "/Users/acanahuati/Downloads/ProjectForAri/ellie_script_csv/Hab_freezing.csv"

OUTPUT_FILE = "dlc_comparison_results.xlsx"

# Columns in the velocity/tracking CSVs to compare
TRACKING_COLS = ["Nose", "Nose.1", "Tail_base", "Tail_base.1",
                 "Tailbase_Speed", "Nose_Speed"]

# ─────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────

def extract_animal_id(filename):
    """
    Pulls the animal ID (e.g. '367_4') from a DLC filename.
    Works on both bare filenames and full paths.
    Pattern: digits, underscore, digits at the start of the basename.
    """
    basename = os.path.basename(filename)
    # Remove common extensions
    for ext in [".csv", ".numbers", "_filtered", "-velocity_data"]:
        basename = basename.replace(ext, "")
    match = re.match(r"(\d+[._]\d+)", basename)
    return match.group(1).replace(".", "_") if match else basename


def style_header(ws, row, n_cols, fill_hex="2F5496"):
    """Bold white text on a dark blue background for header rows."""
    fill = PatternFill("solid", start_color=fill_hex, end_color=fill_hex)
    font = Font(bold=True, color="FFFFFF", name="Arial")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def style_diff_cell(cell, value):
    """Colour-code difference cells: red if |diff| > 0, grey if 0."""
    if pd.isna(value):
        return
    if value == 0:
        cell.font = Font(color="808080", name="Arial")
    else:
        cell.font = Font(color="C00000", bold=True, name="Arial")


def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


# ─────────────────────────────────────────────────────────────
#  BEHAVIOUR SUMMARY COMPARISON
# ─────────────────────────────────────────────────────────────

EPOCH_COLS = ["Habituation", "CS1", "ITI1", "CS2", "ITI2", "CS3", "Post-CS"]

def load_behaviour_csv(path):
    """Load a behaviour summary CSV, extract animal ID from the Number column."""
    df = pd.read_csv(path)
    df["animal_id"] = df["Number"].apply(extract_animal_id)
    # If there are duplicate animal IDs (filtered vs unfiltered), keep _filtered
    df["is_filtered"] = df["Number"].str.contains("_filtered")
    df = df.sort_values("is_filtered", ascending=False).drop_duplicates("animal_id")
    df = df.set_index("animal_id")[EPOCH_COLS]
    return df


def build_behaviour_sheet(ws, my_df, her_df, label):
    """Write a side-by-side comparison sheet for one behaviour metric."""
    animals = sorted(set(my_df.index) | set(her_df.index))

    # Header rows
    ws.append([""] + [label] * len(EPOCH_COLS))
    ws.append(["Animal ID"] +
               [f"Mine\n{e}"     for e in EPOCH_COLS] +
               [f"Hers\n{e}"     for e in EPOCH_COLS] +
               [f"Diff\n{e}"     for e in EPOCH_COLS])
    style_header(ws, 1, 1 + 3 * len(EPOCH_COLS))
    style_header(ws, 2, 1 + 3 * len(EPOCH_COLS))

    for animal in animals:
        my_row  = my_df.loc[animal]  if animal in my_df.index  else pd.Series([np.nan]*len(EPOCH_COLS), index=EPOCH_COLS)
        her_row = her_df.loc[animal] if animal in her_df.index else pd.Series([np.nan]*len(EPOCH_COLS), index=EPOCH_COLS)
        diff    = my_row - her_row

        data_row = [animal] + list(my_row) + list(her_row) + list(diff)
        ws.append(data_row)

        # Colour the diff cells
        excel_row = ws.max_row
        for i, val in enumerate(diff):
            col = 1 + 2 * len(EPOCH_COLS) + 1 + i   # offset to diff columns
            style_diff_cell(ws.cell(excel_row, col), val)

    # Apply font/border to data rows
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.font = cell.font.copy(name="Arial") if cell.font else Font(name="Arial")
            cell.border = thin_border()

    autofit(ws)
    return _behaviour_stats(my_df, her_df, label)


def _behaviour_stats(my_df, her_df, label):
    """Return a dict of per-epoch stats for the summary sheet."""
    animals = sorted(set(my_df.index) & set(her_df.index))
    rows = []
    for epoch in EPOCH_COLS:
        mine = my_df.loc[animals, epoch].values.astype(float)
        hers = her_df.loc[animals, epoch].values.astype(float)
        mask = ~(np.isnan(mine) | np.isnan(hers))
        if mask.sum() < 2:
            corr = mae = mean_diff = np.nan
        else:
            corr      = np.corrcoef(mine[mask], hers[mask])[0, 1]
            mae       = np.mean(np.abs(mine[mask] - hers[mask]))
            mean_diff = np.mean(mine[mask] - hers[mask])
        rows.append({"Metric": label, "Epoch": epoch,
                     "N_animals": mask.sum(),
                     "Mean_Diff (mine-hers)": round(mean_diff, 3),
                     "MAE": round(mae, 3),
                     "Pearson_r": round(corr, 3)})
    return rows


# ─────────────────────────────────────────────────────────────
#  TRACKING CSV COMPARISON
# ─────────────────────────────────────────────────────────────

def match_tracking_files(my_folder, her_folder):
    """Return a list of (animal_id, my_path, her_path) for matched CSVs."""
    my_files  = {extract_animal_id(f): f for f in glob.glob(os.path.join(my_folder,  "*.csv"))}
    her_files = {extract_animal_id(f): f for f in glob.glob(os.path.join(her_folder, "*.csv"))}

    matched, only_mine, only_hers = [], [], []
    all_ids = sorted(set(my_files) | set(her_files))
    for aid in all_ids:
        if aid in my_files and aid in her_files:
            matched.append((aid, my_files[aid], her_files[aid]))
        elif aid in my_files:
            only_mine.append(aid)
        else:
            only_hers.append(aid)

    if only_mine:
        print(f"[INFO] Only in YOUR folder (no match): {only_mine}")
    if only_hers:
        print(f"[INFO] Only in OVERSEER folder (no match): {only_hers}")
    return matched


def load_tracking_csv(path):
    """Load a DLC velocity/tracking CSV. Handles the 3-row header if present."""
    df = pd.read_csv(path)
    # Ensure Second column is present for alignment
    if "Second" not in df.columns and df.columns[0] != "Second":
        df = df.rename(columns={df.columns[0]: "Second"})
    df["Second"] = pd.to_numeric(df["Second"], errors="coerce")
    df = df.dropna(subset=["Second"]).reset_index(drop=True)
    return df


def build_tracking_sheet(ws, my_df, her_df, animal_id):
    """Write frame-by-frame comparison for one animal."""
    cols_present = [c for c in TRACKING_COLS if c in my_df.columns and c in her_df.columns]

    merged = my_df[["Second"] + cols_present].merge(
        her_df[["Second"] + cols_present],
        on="Second", suffixes=("_mine", "_hers"), how="outer"
    ).sort_values("Second").reset_index(drop=True)

    # Build header
    header = ["Second"]
    for c in cols_present:
        header += [f"{c}_mine", f"{c}_hers", f"{c}_diff"]
    ws.append(header)
    style_header(ws, 1, len(header))

    for _, row in merged.iterrows():
        data = [row["Second"]]
        for c in cols_present:
            mine = row.get(f"{c}_mine", np.nan)
            hers = row.get(f"{c}_hers", np.nan)
            diff = (mine - hers) if not (pd.isna(mine) or pd.isna(hers)) else np.nan
            data += [round(mine, 4) if not pd.isna(mine) else "",
                     round(hers, 4) if not pd.isna(hers) else "",
                     round(diff, 4) if not pd.isna(diff) else ""]
        ws.append(data)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name="Arial", size=9)
            cell.border = thin_border()

    autofit(ws)
    return _tracking_stats(merged, cols_present, animal_id)


def _tracking_stats(merged, cols, animal_id):
    rows = []
    for c in cols:
        mine = pd.to_numeric(merged.get(f"{c}_mine"), errors="coerce").values
        hers = pd.to_numeric(merged.get(f"{c}_hers"), errors="coerce").values
        mask = ~(np.isnan(mine) | np.isnan(hers))
        if mask.sum() < 2:
            corr = mae = mean_diff = np.nan
        else:
            corr      = np.corrcoef(mine[mask], hers[mask])[0, 1]
            mae       = np.mean(np.abs(mine[mask] - hers[mask]))
            mean_diff = np.mean(mine[mask] - hers[mask])
        rows.append({"Animal": animal_id, "Column": c,
                     "N_frames": int(mask.sum()),
                     "Mean_Diff (mine-hers)": round(mean_diff, 4),
                     "MAE": round(mae, 4),
                     "Pearson_r": round(corr, 4)})
    return rows


# ─────────────────────────────────────────────────────────────
#  SUMMARY SHEET
# ─────────────────────────────────────────────────────────────

def build_summary_sheet(ws, all_stats):
    df = pd.DataFrame(all_stats)
    if df.empty:
        ws.append(["No data available"])
        return

    ws.append(list(df.columns))
    style_header(ws, 1, len(df.columns))

    for _, row in df.iterrows():
        ws.append(list(row))

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in r:
            cell.font = Font(name="Arial")
            cell.border = thin_border()

    autofit(ws)


# ─────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────

def main():
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet
    all_stats = []

    # ── 1. Behaviour summaries ────────────────────────────────
    behaviour_pairs = [
        (MY_SCAN_CSV,    OVERSEER_SCAN_CSV,    "Scanning"),
        (MY_FREEZE_CSV,  OVERSEER_FREEZE_CSV,  "Freezing"),
    ]

    for my_path, her_path, label in behaviour_pairs:
        if not (os.path.exists(my_path) and os.path.exists(her_path)):
            print(f"[SKIP] {label}: one or both files not found ({my_path}, {her_path})")
            continue
        print(f"[INFO] Comparing behaviour: {label}")
        my_df  = load_behaviour_csv(my_path)
        her_df = load_behaviour_csv(her_path)
        ws = wb.create_sheet(f"{label}_Comparison")
        stats = build_behaviour_sheet(ws, my_df, her_df, label)
        all_stats.extend(stats)

    # ── 2. Tracking CSVs ─────────────────────────────────────
    if os.path.isdir(MY_FOLDER) and os.path.isdir(OVERSEER_FOLDER):
        matched = match_tracking_files(MY_FOLDER, OVERSEER_FOLDER)
        for animal_id, my_path, her_path in matched:
            print(f"[INFO] Comparing tracking: {animal_id}")
            my_df  = load_tracking_csv(my_path)
            her_df = load_tracking_csv(her_path)
            # Truncate sheet name to 31 chars (Excel limit)
            sheet_name = f"Track_{animal_id}"[:31]
            ws = wb.create_sheet(sheet_name)
            stats = build_tracking_sheet(ws, my_df, her_df, animal_id)
            all_stats.extend(stats)
    else:
        print(f"[SKIP] Tracking folders not found. Check MY_FOLDER / OVERSEER_FOLDER in CONFIG.")

    # ── 3. Summary sheet ─────────────────────────────────────
    ws_summary = wb.create_sheet("Summary", 0)   # insert at front
    build_summary_sheet(ws_summary, all_stats)

    wb.save(OUTPUT_FILE)
    print(f"\n✓ Done! Results saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
